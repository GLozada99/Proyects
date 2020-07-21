from bs4 import BeautifulSoup #for html parsing
from urllib.request  import urlopen #for opening the url
from re import sub #for correcting the text
import openpyxl #opening and editing excel files
from yagmail import SMTP #sending mails
from sys import exit #for exiting the script

def correction(s):
    s = sub(r'\. ','.',s)
    s = sub(r'\.','. ',s)#replacing every period with period and space
    s = sub(r'\?','? ',s)#replacing every ? with ? and space
    return s

def get_text_parsed(url):
    url_open = urlopen(url)
    text = url_open.read()
    soup = BeautifulSoup(text,'html.parser')
    return soup

def fill_dict(soup,data_dict):
    #Getting needed data from page
    data_dict['Phrase'] = '\n' + soup.findAll("div",{"class":"field-item even"})[1].text + '\n'  
    data_dict['Definition'] = '\n' + soup.findAll("div",{"class":"field-item even"})[2].text + '\n'  
    example = soup.findAll("div",{"id":"bootstrap-panel-2-body"})[0].text
    data_dict['Example'] = correction(example)
    return data_dict

def create_load_workbook(filename):
    #Loading workbook and getting first spreadsheet
    #If the file doesn't exists, it is created
    try:
        workbook = openpyxl.load_workbook(filename = filename)#load workbook with the name
        sheet = workbook['Phrases']
    except:
        workbook = openpyxl.Workbook()
        elim = workbook['Sheet']
        workbook.remove_sheet(elim)
        sheet = workbook.create_sheet('Phrases')
        for i in range(3):
            print(1,i+1,headers[i])
            sheet.cell(1,i+1).value = headers[i]
            sheet.cell(1,i+1).font = openpyxl.styles.Font(bold=True) 
    return (workbook,sheet)

def check_duplicates(sheet,phrase):
    last_cell = sheet.cell(sheet.max_row,1)#getting last phrase cell that has a value    
    if last_cell.value == phrase: 
        print('The phrase was already sent today')
        exit()

def get_emails(file):
    fil = open(file,'r')
    name_ls = []
    mail_ls = []
    dicti = {'Names':name_ls,'Mails':mail_ls}
    for line in fil:
        stripped_line = line.strip().split(" ")
        dicti['Names'].append(stripped_line[0])
        dicti['Mails'].append(stripped_line[1])
    return dicti

def send_mails(emitter, password, contacts, content):
    yag = SMTP(emitter,password)
    phr = content['Phrase']
    des = content['Definition']
    exa = content['Example']
    body = 'Phrase: {}\nDefinition: {}\nExample: {}'.format(phr,des,exa) 
    for name,reciever in zip(contacts['Names'],contacts['Mails']):
        yag.send(
            to=reciever,
            subject="Hola {}, esta es la frase de hoy".format(name),
            contents=body, 
        )
    yag.close()
    
            
#Connecting to url and constructing dict for the data
my_url = 'https://www.ihbristol.com/english-phrases'
soup = get_text_parsed(my_url)
headers = ['Phrase','Definition','Example']
data_dict = {headers[0]:'',headers[1]:'',headers[2]:''}
data_dict = fill_dict(soup,data_dict)

#Opening or loading excel document, checking for duplicates and ending the script if they exist
excel_filename = '/home/gustavolozada/Documents/PyPrograming/ExcelDocs/PhrasesOfTheDay.xlsx'
workbook, sheet = create_load_workbook(excel_filename) 
check_duplicates(sheet,data_dict['Phrase'])

#Iterating through the dict to write on spreadsheet
blank_row = sheet.max_row
for head,col in zip(headers,range(3)):
    print(head + ": "+ data_dict[head])

    #The row and column must be at least 1
    cell = (blank_row + 1, col + 1,openpyxl.utils.get_column_letter(col + 1)) #row,column,column_letter

    sheet.cell(cell[0],cell[1]).value = data_dict[head] 
    
    sheet.cell(cell[0],cell[1]).alignment = openpyxl.styles.Alignment\
    (wrapText=True, horizontal='center',vertical='center')
    
    sheet.column_dimensions[cell[2]].width = float(20 + 15*col)

sheet.row_dimensions[blank_row+1].height = float(60)#last column (description) must have bigger width


#Save and close the excel file
workbook.save(excel_filename)
workbook.close()

#Get and open contacts file
contacts_file = '/home/gustavolozada/Documents/PyPrograming/Contacts.txt'
contacts = get_emails(contacts_file)

#Sending mails to contacts from the emitter
emitter = '''EMAIL FROM WHICH THE MAIL IS GOING TO BE SEND (MUST BE GMAIL)'''
password = '''PASSWORD FOR THAT EMAIL'''
send_mails(emitter,password,contacts,data_dict) 










